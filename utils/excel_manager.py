from typing import Any, Generator, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import logger


class ExcelManager:
    """Handle reading and writing of the KPI workbook.

    This class relies on :mod:`openpyxl` so that individual cells can be
    updated and the file saved incrementally.  The expected spreadsheet
    structure is:

    * first row is a header row containing at least the columns
      "Key Parameters" and "KPIs" (order may vary)
    * subsequent rows contain a JSON string in the "Key Parameters" column
      and either an empty cell or a JSON string in the "KPIs" column.
    """

    def __init__(self, path: str) -> None:
        self.path = path
        self.workbook = load_workbook(self.path)
        logger.info(f"loaded workbook '{self.path}' with sheets: {self.workbook.sheetnames}")

    def iterate_sheets(self) -> Generator[Worksheet, None, None]:
        """Yield every worksheet in the workbook."""
        for sheet in self.workbook.worksheets:
            yield sheet

    def find_columns(self, sheet: Worksheet) -> Tuple[int, int]:
        """Locate the key and KPI column indices for a sheet.

        Raises :class:`ValueError` if either header is missing.
        """
        key_idx: Optional[int] = None
        kpi_idx: Optional[int] = None
        for cell in sheet[1]:
            if cell.value is None:
                continue
            header = str(cell.value).strip()
            if header == "Key Parameters":
                key_idx = cell.column
            elif header == "KPIs":
                kpi_idx = cell.column
        if key_idx is None or kpi_idx is None:
            raise ValueError(f"sheet '{sheet.title}' missing required headers")
        return key_idx, kpi_idx

    def iter_sheet_rows(self, sheet: Worksheet) -> Generator[Tuple[int, Any, Any], None, None]:
        """Yield ``(row_index, key_cell, kpi_cell)`` for every data row.

        ``row_index`` starts at 2 (assuming header in row 1) and continues to
        ``sheet.max_row``.
        """
        key_col, kpi_col = self.find_columns(sheet)
        for row in range(2, sheet.max_row + 1):
            key_cell = sheet.cell(row=row, column=key_col)
            kpi_cell = sheet.cell(row=row, column=kpi_col)
            yield row, key_cell, kpi_cell

    def save(self) -> None:
        """Persist changes to disk."""
        try:
            self.workbook.save(self.path)
            logger.debug(f"workbook saved to '{self.path}'")
        except Exception as e:
            logger.error(f"error saving workbook '{self.path}': {e}")
